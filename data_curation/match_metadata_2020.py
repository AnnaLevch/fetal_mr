import openpyxl
import os
import pandas as pd


def case_id_from_name(case_folder):
    splitted = case_folder.split('_')
    case_id = splitted[0] + '_' + splitted[1]
    return case_id[3:], splitted[2][2:]


if __name__ == "__main__":
    metadata_dir = 'C:/Bella/data/NewCases2020/'
    cases_lst_filename = 'trufi_cases.csv'
    ids_matching_filename = 'metadata_all.xlsx'
    metadata_filename = 'metadataFor2020.xlsx'
    save_filename = 'matched_metadata_2020.xlsx'

    trufi_cases = pd.read_csv(os.path.join(metadata_dir, cases_lst_filename))['0'].to_list()
    matching_data = pd.read_excel(os.path.join(metadata_dir,ids_matching_filename), engine='openpyxl')
    metadata = openpyxl.load_workbook(os.path.join(metadata_dir,metadata_filename))
    medatada_worksheet = metadata.active

    id_filename_dict = {}
    for case in trufi_cases:
        filename_id, series_id = case_id_from_name(case)
        case_matching = matching_data.loc[matching_data['filename'].str.contains(filename_id)]
        id = case_matching['ID'].values[0]
        if id in id_filename_dict:
            print('id ' + str(id) + ' is already in dictionary. Current filename id is: '
                  + filename_id + ', in dictionary: ' + id_filename_dict[id][0])
            print('prev series: ' + id_filename_dict[id][1] + ', current series id: ' + series_id)

        id_filename_dict[id] = [filename_id, series_id]

    #append filename id for matched rows and remove unmatched rows
    medatada_worksheet.insert_cols(medatada_worksheet.max_column+1,1)
    medatada_worksheet.cell(column=medatada_worksheet.max_column, row=1).value='filename'
    i=2
    for row in medatada_worksheet.iter_rows():
        id = medatada_worksheet.cell(column=4, row=i).value
        if id in id_filename_dict:
            print('id: '+ id + ' is in dict' )
            medatada_worksheet.cell(column=medatada_worksheet.max_column, row=i).value=id_filename_dict[id][0]
            i += 1
        else:
            medatada_worksheet.delete_rows(i)


    metadata.save(os.path.join(metadata_dir, save_filename))